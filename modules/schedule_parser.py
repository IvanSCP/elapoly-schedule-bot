from openpyxl import load_workbook
from datetime import datetime, timedelta
from typing import Optional
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

months = {
    1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля',
    5: 'мая', 6: 'июня', 7: 'июля', 8: 'августа',
    9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
}

async def parse_schedule_for_date(schedule_file: str, config: dict, target_date: datetime) -> Optional[str]:
    """
    Парсит расписание для указанной даты из Excel-файла.
    
    Args:
        schedule_file (str): Путь к файлу с расписанием
        config (dict): Конфигурация с настройками парсинга
        target_date (datetime): День, на который нужно получить расписание
        
    Returns:
        Optional[str]: Отформатированное расписание или None в случае ошибки
    """
    try:
        # Проверяем существование файла
        if not Path(schedule_file).exists():
            logger.error(config['logger_messages']['schedule_file_not_found'])
            return None

        # Загружаем книгу и выбираем нужный лист
        wb = load_workbook(schedule_file)
        ws = wb[config['schedule_parser']['target_sheet']]

        # Форматируем строку для поиска даты
        date_str = f"{target_date.day} {months[target_date.month]}".upper()
        
        date_col = config['schedule_parser']['date_column']
        time_col = config['schedule_parser']['time_column']
        group_name = config['schedule_parser']['group_name']
        rows_to_fetch = config['schedule_parser']['rows_to_fetch']

        # Ищем строку с нужной датой
        date_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws[f'{date_col}{row}'].value
            if cell_value and isinstance(cell_value, str) and date_str in cell_value.upper():
                date_row = row
                break

        if not date_row:
            return None

        # Находим строку с группами (обычно через 4 строки после даты)
        groups_row = date_row + 4

        # Динамически ищем колонку с нужной группой
        group_column = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=groups_row, column=col).value
            if cell_value and group_name in str(cell_value):
                group_column = col
                break

        if not group_column:
            return None

        # Собираем данные о расписании
        schedule_lines = []
        current_time = None

        # Проходим по 16 строкам с расписанием "под группой" (если в файле их больше, то измените конфиг!)
        for i in range(1, rows_to_fetch + 1):
            row_index = groups_row + i
            
            # Проверяем ячейку времени
            time_value = ws[f'{time_col}{row_index}'].value
            
            # Если нашли новое время, добавляем его в расписание
            if time_value and time_value != current_time:
                if current_time is not None:
                    schedule_lines.append("")  # Пустая строка между временными блоками
                current_time = time_value
                schedule_lines.append(current_time)
            
            # Добавляем данные из колонки группы (если есть)
            group_value = ws.cell(row=row_index, column=group_column).value
            if group_value:
                schedule_lines.append(str(group_value))

        # Форматируем и возвращаем результат
        if schedule_lines:
            schedule_text = "\n".join(schedule_lines)
            return f"{schedule_text}" # Расписание на {date_str.lower()}:\n
        'else:' # Думаю, что эта проверка лишняя
        #    return f"На {date_str.lower()} занятий нет"
    
    except Exception as e:
        logger.error(config['logger_messages']['schedule_parser_error'].format(e=e))
        return None

async def parse_schedule_for_today(schedule_file: str, config: dict) -> Optional[str]:
    """Парсит расписание на сегодня"""
    today = (datetime.now())
    return await parse_schedule_for_date(schedule_file, config, today)

async def parse_schedule_for_tomorrow(schedule_file: str, config: dict) -> Optional[str]:
    """Парсит расписание на завтра"""
    tomorrow = (datetime.now() + timedelta(days=1))
    return await parse_schedule_for_date(schedule_file, config, tomorrow)